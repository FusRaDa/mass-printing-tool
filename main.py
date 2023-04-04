import sys
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
import os


def generate_excel_file():
    warning_title = "File Execution Failed"

    location_val = location.get()
    item_number_val = item_number.get()
    lot_val = lot.get()
    quantity_val = quantity.get()
    number_of_labels_val = number_of_labels.get()

    if location_val.__contains__(" ") or len(location_val) == 0 \
            or item_number_val.__contains__(" ") or len(item_number_val) == 0 \
            or lot_val.__contains__(" ") or len(lot_val) == 0 \
            or quantity_val.__contains__(" ") or len(quantity_val) == 0 \
            or number_of_labels_val.__contains__(" ") or len(number_of_labels_val) == 0:
        messagebox.showerror(warning_title, "Fill all inputs, values must not contain spaces.")
        return

    try:
        float(quantity_val)
    except ValueError:
        messagebox.showerror(warning_title, "Quantity must only be a numbers.")
        return

    try:
        int(number_of_labels_val)
    except ValueError:
        messagebox.showerror(warning_title, "Number of labels must only be a whole number.")
        return

    try:
        row_data = [location_val, item_number_val, lot_val, quantity_val]

        wb = Workbook()
        ws = wb.active

        ws['A1'] = "Location"
        ws['B1'] = "Item #"
        ws['C1'] = "Lot"
        ws['D1'] = "Quantity"

        last_cell_row = int(number_of_labels_val)
        last_cell_col = len(row_data)

        for row in range(2, last_cell_row + 2):
            for col in range(1, last_cell_col + 1):
                ws.cell(row=row, column=col).value = row_data[col - 1]

        wb.save("mass_printing_tool_output.xlsx")
    except PermissionError:
        messagebox.showerror(warning_title, "Please close out of file before running program again")
        return

    os.system('start excel.exe "%s\\mass_printing_tool_output.xlsx"' % (sys.path[0]))


# run ui
window = tk.Tk()
window.title("Mass Printing Generator")

tk.Label(window, text="Location").grid(row=0, column=1)
tk.Label(window, text="Item #").grid(row=0, column=2)
tk.Label(window, text="Lot").grid(row=0, column=3)
tk.Label(window, text="Quantity").grid(row=0, column=4)

tk.Label(window, text="Number of Labels").grid(row=2, column=2, pady=(15, 1))

tk.Button(window, text="Generate Excel File", command=generate_excel_file).grid(row=3, column=3)

location = tk.Entry(window)
item_number = tk.Entry(window)
lot = tk.Entry(window)
quantity = tk.Entry(window)

number_of_labels = tk.Entry(window)

location.grid(row=1, column=1)
item_number.grid(row=1, column=2)
lot.grid(row=1, column=3)
quantity.grid(row=1, column=4)

number_of_labels.grid(row=3, column=2)

window.mainloop()
